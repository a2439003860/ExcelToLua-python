#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel转Lua转换器
功能：
1. 转换当前目录下excel文件夹内的xlsx文件为lua文件
2. 跳过文件名包含$或#的文件
3. 跳过首单元格以/或?开头的行
4. 支持数据类型标记和数组类型（用[]表示一维数组，用{}表示二维数组）
5. 检测并提醒重复的ID
6. 跳过第三行参数为空的列
"""

import os
import sys
import re
from typing import Dict, List, Any, Tuple, Optional
import warnings

# 尝试导入openpyxl，如果未安装则提示
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 未找到openpyxl库。请使用以下命令安装：")
    print("pip install openpyxl")
    sys.exit(1)


class ExcelToLuaConverter:
    def __init__(self):
        self.excel_dir = os.path.join(os.path.dirname(__file__), "excel")
        self.lua_dir = os.path.join(os.path.dirname(__file__), "lua")
        self.duplicate_ids = {}  # 存储重复ID信息
        self.skipped_columns = {}  # 存储跳过的列信息[6](@ref)
        
    def ensure_dirs(self):
        """确保必要的目录存在"""
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
            print(f"创建excel目录: {self.excel_dir}")
            
        if not os.path.exists(self.lua_dir):
            os.makedirs(self.lua_dir)
            print(f"创建lua目录: {self.lua_dir}")
    
    def should_skip_file(self, filename: str) -> bool:
        """判断是否需要跳过文件"""
        # 跳过包含$或#的文件
        if '$' in filename or '#' in filename:
            return True
        
        # 只处理.xlsx文件
        if not filename.lower().endswith('.xlsx'):
            return True
            
        return False
    
    def should_skip_row(self, first_cell_value) -> bool:
        """判断是否需要跳过的行"""
        if first_cell_value is None:
            return True
            
        cell_str = str(first_cell_value).strip()
        if cell_str.startswith('/') or cell_str.startswith('?'):
            return True
            
        return False
    
    def is_column_empty(self, key_cell_value) -> bool:
        """判断列是否需要跳过（第三行参数为空）"""
        if key_cell_value is None:
            return True
            
        key_str = str(key_cell_value).strip()
        return len(key_str) == 0
    
    def parse_data_type(self, type_str: str) -> Tuple[str, int]:
        """解析数据类型，返回(类型, 数组维度)
        0: 非数组, 1: 一维数组, 2: 二维数组
        """
        if not type_str:
            return "string", 0
            
        type_str = str(type_str).strip().lower()
        array_dimension = 0
        
        # 检查是否为二维数组类型 {type}
        match_2d = re.match(r'\{(.*?)\}', type_str)
        if match_2d:
            type_str = match_2d.group(1)
            array_dimension = 2
        else:
            # 检查是否为一维数组类型 [type]
            match_1d = re.match(r'\[(.*?)\]', type_str)
            if match_1d:
                type_str = match_1d.group(1)
                array_dimension = 1
        
        # 标准化类型名
        if type_str in ['int', 'integer', 'number']:
            return "int", array_dimension
        elif type_str in ['float', 'double', 'decimal']:
            return "float", array_dimension
        elif type_str in ['bool', 'boolean']:
            return "bool", array_dimension
        elif type_str in ['str', 'string']:
            return "string", array_dimension
        else:
            # 默认为string类型
            return "string", array_dimension
    
    def convert_value(self, value, data_type: str, array_dimension: int):
        """根据数据类型和数组维度转换值"""
        if array_dimension == 0:
            # 非数组类型
            return self._convert_single_value(value, data_type)
        elif array_dimension == 1:
            # 一维数组：使用逗号分割
            return self._convert_1d_array(value, data_type)
        elif array_dimension == 2:
            # 二维数组：一维用逗号分割，二维用分号分割
            return self._convert_2d_array(value, data_type)
        else:
            return self._convert_single_value(value, data_type)
    
    def _convert_1d_array(self, value, data_type: str) -> List:
        """转换一维数组（使用逗号分割）"""
        if value is None:
            return []
        
        str_value = str(value).strip()
        if not str_value:
            return []
        
        # 按逗号分割，过滤空值
        parts = [part.strip() for part in str_value.split(',') if part.strip()]
        
        # 转换每个元素
        converted_parts = []
        for part in parts:
            converted_parts.append(self._convert_single_value(part, data_type))
        
        return converted_parts
    
    def _convert_2d_array(self, value, data_type: str) -> List[List]:
        """转换二维数组（一维逗号分割，二维分号分割）"""
        if value is None:
            return []
        
        str_value = str(value).strip()
        if not str_value:
            return []
        
        # 按分号分割二维，过滤空值
        # 支持数据尾部无论是否用分号结尾都能正确转换
        lines = [line.strip() for line in str_value.split(';') if line.strip()]
        
        converted_array = []
        for line in lines:
            # 按逗号分割一维，过滤空值
            parts = [part.strip() for part in line.split(',') if part.strip()]
            
            # 转换每个元素
            converted_line = []
            for part in parts:
                converted_line.append(self._convert_single_value(part, data_type))
            
            if converted_line:  # 只添加非空行
                converted_array.append(converted_line)
        
        return converted_array
    
    def _convert_single_value(self, value, data_type: str):
        """转换单个值"""
        if value is None:
            if data_type == "string":
                return ""
            elif data_type == "bool":
                return False
            else:
                return 0
        
        str_value = str(value).strip()
        
        if data_type == "int":
            try:
                # 尝试转换为整数
                if '.' in str_value:
                    return int(float(str_value))
                return int(str_value)
            except:
                # 转换失败，尝试提取数字
                numbers = re.findall(r'-?\d+', str_value)
                return int(numbers[0]) if numbers else 0
                
        elif data_type == "float":
            try:
                return float(str_value)
            except:
                numbers = re.findall(r'-?\d+\.?\d*', str_value)
                return float(numbers[0]) if numbers else 0.0
                
        elif data_type == "bool":
            str_lower = str_value.lower()
            if str_lower in ['true', 'yes', 'y', '1', '是', '真']:
                return True
            elif str_lower in ['false', 'no', 'n', '0', '否', '假']:
                return False
            else:
                # 尝试数值转换
                try:
                    num = float(str_value)
                    return bool(num)
                except:
                    return False
                    
        elif data_type == "string":
            # 处理字符串，转义特殊字符
            str_value = str_value.replace('\\', '\\\\')
            str_value = str_value.replace('"', '\\"')
            str_value = str_value.replace('\n', '\\n')
            str_value = str_value.replace('\r', '\\r')
            return str_value
            
        else:
            return str_value
    
    def convert_excel_to_lua(self, excel_path: str) -> bool:
        """转换单个Excel文件为Lua"""
        try:
            # 加载工作簿
            wb = load_workbook(filename=excel_path, data_only=True)
            ws = wb.active  # 获取第一个工作表
            
            print(f"处理文件: {os.path.basename(excel_path)}")
            
            # 获取列数
            max_column = ws.max_column
            max_row = ws.max_row
            
            if max_row < 3:
                print(f"  警告: 文件行数不足3行，跳过")
                return False
            
            # 读取数据类型行（第2行）和key行（第3行）
            data_types = []
            keys = []
            valid_columns = []  # 记录有效列的索引[6](@ref)
            is_id_col = -1  # ID列的索引
            
            for col in range(1, max_column + 1):
                # 检查第三行参数是否为空[6](@ref)
                key_cell = ws.cell(row=3, column=col).value
                if self.is_column_empty(key_cell):
                    continue  # 跳过该列[7](@ref)
                
                # 记录有效列
                valid_columns.append(col)
                
                # 数据类型
                type_cell = ws.cell(row=2, column=col).value
                data_type, array_dimension = self.parse_data_type(type_cell)
                data_types.append((data_type, array_dimension))
                
                # key
                key = str(key_cell).strip() if key_cell else f"col_{col}"
                keys.append(key)
                
                # 检查是否为ID列
                if key.lower() == 'id':
                    is_id_col = len(keys) - 1  # 更新为有效列中的索引
            
            # 记录跳过的列信息[6](@ref)
            filename = os.path.basename(excel_path)
            skipped_count = max_column - len(valid_columns)
            if skipped_count > 0:
                self.skipped_columns[filename] = skipped_count
            
            if len(valid_columns) == 0:
                print(f"  警告: 没有有效的列，跳过文件")
                return False
            
            # 收集数据行
            data_rows = []
            id_values = []  # 存储ID值用于重复检查
            
            for row in range(4, max_row + 1):
                # 检查第一列是否需要跳过
                first_cell = ws.cell(row=row, column=1).value
                if self.should_skip_row(first_cell):
                    continue
                
                # 读取该行数据（只读取有效列）
                row_data = {}
                for idx, col in enumerate(valid_columns):
                    key = keys[idx]
                    data_type, array_dimension = data_types[idx]
                    cell_value = ws.cell(row=row, column=col).value
                    
                    # 转换值
                    converted_value = self.convert_value(cell_value, data_type, array_dimension)
                    row_data[key] = converted_value
                
                # 如果是有效行，添加到数据列表
                if row_data:
                    data_rows.append(row_data)
                    
                    # 记录ID值用于重复检查
                    if is_id_col >= 0 and is_id_col < len(keys):
                        id_key = keys[is_id_col]
                        if id_key in row_data:
                            id_value = row_data[id_key]
                            id_values.append(id_value)
            
            # 检查重复ID
            if is_id_col >= 0:
                id_counts = {}
                for id_val in id_values:
                    id_counts[id_val] = id_counts.get(id_val, 0) + 1
                
                duplicate_ids = [id_val for id_val, count in id_counts.items() if count > 1]
                if duplicate_ids:
                    if filename not in self.duplicate_ids:
                        self.duplicate_ids[filename] = []
                    self.duplicate_ids[filename].extend(duplicate_ids)
            
            # 生成Lua文件内容
            lua_content = self.generate_lua_content(data_rows, keys, os.path.basename(excel_path))
            
            # 保存Lua文件
            excel_name = os.path.basename(excel_path)
            lua_filename = os.path.splitext(excel_name)[0] + '.lua'
            lua_path = os.path.join(self.lua_dir, lua_filename)
            
            with open(lua_path, 'w', encoding='utf-8') as f:
                f.write(lua_content)
            
            print(f"  成功: 生成 {lua_filename} ({len(data_rows)}行数据, {len(valid_columns)}列数据)")
            return True
            
        except Exception as e:
            print(f"  错误: 处理文件失败 - {str(e)}")
            return False
    
    def generate_lua_content(self, data_rows: List[Dict], keys: List[str], excel_filename: str) -> str:
        """生成Lua文件内容"""
        # 使用文件名（去掉扩展名）作为表名
        table_name = os.path.splitext(excel_filename)[0]
        
        lua_lines = [f"-- 自动生成的Lua配置文件", f"-- 源文件: {excel_filename}", ""]
        
        # 创建命名表
        lua_lines.append(f"{table_name} = {{")
        
        for i, row in enumerate(data_rows):
            # 使用自增索引作为key值，从1开始
            key_line = f"    [{i+1}] = {{"
            lua_lines.append(key_line)
            
            # 添加数据字段
            field_lines = []
            for key_name, value in row.items():
                if isinstance(value, bool):
                    value_str = "true" if value else "false"
                elif isinstance(value, (int, float)):
                    value_str = str(value)
                elif isinstance(value, str):
                    value_str = f'"{value}"'
                elif isinstance(value, list):
                    # 判断是否为二维数组（嵌套列表）
                    if value and isinstance(value[0], list):
                        # 二维数组处理
                        array_lines = []
                        for sub_array in value:
                            sub_items = []
                            for item in sub_array:
                                if isinstance(item, bool):
                                    sub_items.append("true" if item else "false")
                                elif isinstance(item, (int, float)):
                                    sub_items.append(str(item))
                                elif isinstance(item, str):
                                    sub_items.append(f'"{item}"')
                                else:
                                    sub_items.append(f'"{str(item)}"')
                            array_lines.append("{" + ", ".join(sub_items) + "}")
                        value_str = "{" + ", ".join(array_lines) + "}"
                    else:
                        # 一维数组处理
                        array_items = []
                        for item in value:
                            if isinstance(item, bool):
                                array_items.append("true" if item else "false")
                            elif isinstance(item, (int, float)):
                                array_items.append(str(item))
                            elif isinstance(item, str):
                                array_items.append(f'"{item}"')
                            else:
                                array_items.append(f'"{str(item)}"')
                        value_str = "{" + ", ".join(array_items) + "}"
                else:
                    value_str = f'"{str(value)}"'
                
                field_lines.append(f"        {key_name} = {value_str}")
            
            lua_lines.append(",\n".join(field_lines))
            lua_lines.append("    },")
        
        lua_lines.append("}")
        lua_lines.append("")
        lua_lines.append(f"return {table_name}")
        
        return "\n".join(lua_lines)
    
    def print_duplicate_warnings(self):
        """打印重复ID的警告信息"""
        if not self.duplicate_ids:
            return
        
        print("\n" + "="*60)
        print("警告: 发现重复的ID值")
        print("="*60)
        
        for filename, ids in self.duplicate_ids.items():
            print(f"\n文件: {filename}")
            for dup_id in set(ids):
                print(f"  重复ID: {dup_id}")
        
        print("\n注意: 请检查上述重复的ID，这可能导致数据覆盖问题！")
        print("="*60)
    
    def print_skipped_columns_info(self):
        """打印跳过的列信息"""
        if not self.skipped_columns:
            return
        
        print("\n" + "="*60)
        print("跳过的列统计（第三行参数为空）")
        print("="*60)
        
        for filename, skipped_count in self.skipped_columns.items():
            print(f"文件: {filename} - 跳过了 {skipped_count} 列")
        
        print("\n注意: 以上列因第三行参数为空而被跳过")
        print("="*60)
    
    def run_conversion(self):
        """运行转换过程"""
        print("Excel转Lua转换器")
        print("-" * 40)
        
        # 确保目录存在
        self.ensure_dirs()
        
        # 获取所有Excel文件
        if not os.path.exists(self.excel_dir):
            print(f"错误: 未找到excel目录: {self.excel_dir}")
            return
        
        excel_files = []
        for filename in os.listdir(self.excel_dir):
            if self.should_skip_file(filename):
                continue
            excel_files.append(os.path.join(self.excel_dir, filename))
        
        if not excel_files:
            print("未找到可转换的Excel文件")
            print(f"请将.xlsx文件放入: {self.excel_dir}")
            return
        
        print(f"找到 {len(excel_files)} 个Excel文件")
        print("-" * 40)
        
        # 转换每个文件
        success_count = 0
        for excel_path in excel_files:
            if self.convert_excel_to_lua(excel_path):
                success_count += 1
        
        # 显示跳过的列信息
        self.print_skipped_columns_info()
        
        # 显示重复ID警告
        self.print_duplicate_warnings()
        
        print("-" * 40)
        print(f"转换完成: {success_count}/{len(excel_files)} 个文件转换成功")
        print(f"Lua文件保存到: {self.lua_dir}")


def main():
    """主函数"""
    converter = ExcelToLuaConverter()
    converter.run_conversion()
    
    # 等待用户按键（在批处理中运行时不等待）
    if sys.platform == "win32" and sys.stdin.isatty():
        input("\n按Enter键退出...")


if __name__ == "__main__":
    main()